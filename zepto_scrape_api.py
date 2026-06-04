"""
Zepto Product Details Scraper — Selenium-based, Multi-Pincode
=============================================================
Reads product links from Grocery_Template.xlsx (sheet: Zepto) and pincodes from
pincodes_to_run.xlsx — both expected to sit in the SAME folder as this script.
Scrapes all products for each pincode and writes a multi-sheet output Excel
(also in the same folder) named Zepto_{YYYY-MM-DD}_{HH-MM-SS}.xlsx.

Mirrors the Blinkit / FK Minutes scrapers: same input-file handling, dated
outputs, per-pincode workbook sheets + Summary + Price_Comparison, resume via a
dated progress JSON, a combined product_data JSON, and a batched OOS fallback.

Zepto specifics (preserved from the unified scraper):
  - A single Selenium Chrome driver is used for the whole run; its delivery
    location is re-set per pincode via the real Zepto location UI
    (Select Location -> enter pincode -> Confirm & Continue).
  - Each product page is scraped with Selenium (_zpt_scrape_page): ProductName,
    Price, MRP, ZeptoSaverPrice, stock, Net Qty -> Weight, and the SSR store_id.
  - Because one driver can switch location, OOS fallback just re-points the same
    driver at each fallback pincode (no new browser per pincode like FK needs).

Output workbook contains:
  - One sheet per pincode: "Tier_Pincode" (e.g. "Metro_400051")
  - "Summary" sheet: flagged zero-price issues
  - "Price_Comparison" sheet: all products with MRP/Price across pincodes

Also writes (next to this script, prefixed with the run date {date}):
  - {date}_zepto_progress.json      — resume state: per-pincode success/errored indices
  - {date}_zepto_product_data.json  — merged scraped data, all pincodes (keyed by pincode)

Resume:
  --resume reuses the MOST RECENT dated progress file. Successful rows are
  skipped; errored rows (invalid/failed scrapes) are re-scraped, as are rows
  never attempted. Fully-completed pincodes are skipped entirely.

OOS fallback (batched / breadth-first):
  Pass 1 scrapes every product at the primary pincode. Pass 2 takes the still
  Out-of-Stock (or failed) products, re-points the driver at each mapped fallback
  pincode IN ORDER (from the 'Map' sheet), and re-checks the whole batch; any
  product In Stock there resolves and drops out. Pincode_Used records the winner.

CLI:
  python zepto_scrape_api.py
  python zepto_scrape_api.py --resume
  python zepto_scrape_api.py --company VMM --delay 2.0 --headless

Dependencies:
  pip install selenium webdriver-manager openpyxl pandas
"""
import re
import copy
import time
import json
import datetime
import traceback
from pathlib import Path
from typing import Any, Optional, Tuple, Union

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# ---------------------------------------------------------------------------
#  CONFIGURATION
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "Grocery_Template.xlsx"
PINCODES_PATH = BASE_DIR / "pincodes_to_run.xlsx"

# Resume-state + merged-data JSON files. These are *base* names; run_multi_pincode_scraper()
# rebinds both to {date}_-prefixed paths at startup, and --resume reuses the most
# recent existing dated progress file.
PROGRESS_FILE = BASE_DIR / "zepto_progress.json"
COMBINED_DATA_PATH = BASE_DIR / "zepto_product_data.json"
SAVE_EVERY = 10          # flush progress every N processed products

TEMPLATE_COLUMNS = [
    "Item",
    "Quantity",
    "Base_Final quantity",   # target qty the closest-backup selection matches against
    "Base_Final unit",
    "Segment",
    "Grocery",
    "National / PL",
    "Product_Link",       # the Zepto product URL
    "Backup_Links",          # alternate variant URLs tried when primary qty mismatches
]

SCRAPED_COLUMNS = [
    "ProductName",
    "Price",
    "MRP",
    "ZeptoSaverPrice",
    "is_out_of_stock",
    "Weight",
    "Store_ID",
    "Final_Qty",
    "Final_unit",
    "Pincode_Used",       # primary pincode, or the mapped fallback that returned In Stock
    "Used_Link",          # Product_Link, or a Backup_Links entry if a backup was a closer qty match
]

# Fields populated by _zpt_scrape_page (the scraper's raw output, before the
# Pincode_Used / Used_Link bookkeeping columns are added). Kept as a dict keyed
# by platform to match the helper lifted verbatim from the unified scraper.
SCRAPED_FIELDS = {
    "zepto": [
        "ProductName", "Price", "MRP", "ZeptoSaverPrice",
        "is_out_of_stock", "Weight", "Store_ID", "Final_Qty", "Final_unit",
    ],
}

# Highlight fills / fonts (shared look with the Blinkit / FK workbooks)
MISMATCH_FILL = PatternFill("solid", fgColor="FFC7CE")
ZERO_PRICE_FILL = PatternFill("solid", fgColor="FFEB9C")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
#  PINCODE NORMALIZATION
# ---------------------------------------------------------------------------

def _clean_pincode_str(val) -> str:
    """Normalize a pincode to a clean string: '400051.0' / 400051.0 -> '400051'.
    Non-numeric values are returned stripped (so genuine junk is still visible)."""
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return ""
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s

def _pincode_int(val):
    """Coerce a pincode to int, tolerating '400051.0' floats-as-strings.
    Returns None if it can't be parsed."""
    try:
        return int(float(str(val).strip()))
    except (TypeError, ValueError):
        return None

# ---------------------------------------------------------------------------
#  WEIGHT PARSING (Zepto) + ID/LINK CLEANING
# ---------------------------------------------------------------------------

def _parse_weight_to_base_zepto(weight_str: Any) -> Tuple[Optional[Union[int, float]], Optional[str]]:
    """Parse Zepto Net Qty strings -> (qty in g/ml, unit).

    Mirrors the dedicated Zepto parser used in the standalone Grocery_Tracker
    review script. Handles multipliers (`6 x 70 g`, `500 g x 2`, `12 × 250 ml`),
    parenthetical net qty (`Pack of 6 (5 kg)`, `10 N (1 kg)` — picks the value
    inside the parens, not the count outside), and simple `1.5 kg` / `750 ml`
    formats. Returns (None, None) when no usable qty is found (vs. the generic
    `_parse_weight_to_base` which falls back to `0.0, "g"`).
    """
    if not weight_str:
        return (None, None)

    s = str(weight_str).strip().lower()
    s = s.replace("×", "x")
    s = re.sub(r"\s+", " ", s)

    def to_base(amount, unit):
        try:
            val = float(amount)
        except Exception:
            return (None, None)
        unit = unit.lower()
        if unit == "g":  return (val, "g")
        if unit == "kg": return (val * 1000, "g")
        if unit == "ml": return (val, "ml")
        if unit in ("l", "lt", "ltr", "liter", "litre"):
            return (val * 1000, "ml")
        return (None, None)

    # 1) Multipliers
    # Pattern A: count x amount unit  (e.g. "6 x 70 g", "12 x 250 ml")
    m = re.search(
        r'(?P<count>\d+(?:\.\d+)?)\s*x\s*(?P<amt>\d+(?:\.\d+)?)\s*(?P<unit>kg|g|l|ml|ltr|lt|liter|litre)',
        s)
    if m:
        count = float(m.group("count")); amt = float(m.group("amt"))
        base_amt, base_unit = to_base(amt, m.group("unit"))
        if base_amt is None: return (None, None)
        total = count * base_amt
        return (int(total) if float(total).is_integer() else total, base_unit)

    # Pattern B: amount unit x count  (e.g. "500 g x 2", "1 L x 6")
    m = re.search(
        r'(?P<amt>\d+(?:\.\d+)?)\s*(?P<unit>kg|g|l|ml|ltr|lt|liter|litre)\s*x\s*(?P<count>\d+(?:\.\d+)?)',
        s)
    if m:
        count = float(m.group("count")); amt = float(m.group("amt"))
        base_amt, base_unit = to_base(amt, m.group("unit"))
        if base_amt is None: return (None, None)
        total = count * base_amt
        return (int(total) if float(total).is_integer() else total, base_unit)

    # 2) Value inside parentheses, e.g. "Pack of 6 (5 kg)", "10 N (1 kg)"
    par = re.search(r"\(([^)]+)\)", s)
    if par:
        inner = par.group(1)
        m = re.search(
            r'(?P<amt>\d+(?:\.\d+)?)\s*(?P<unit>kg|g|l|ml|ltr|lt|liter|litre)',
            inner)
        if m:
            base_amt, base_unit = to_base(m.group("amt"), m.group("unit"))
            if base_amt is None: return (None, None)
            return (int(base_amt) if float(base_amt).is_integer() else base_amt, base_unit)

    # 3) Simple format, e.g. "500 g", "1.5 kg", "750 ml"
    m = re.search(
        r'(?P<amt>\d+(?:\.\d+)?)\s*(?P<unit>kg|g|l|ml|ltr|lt|liter|litre)',
        s)
    if m:
        base_amt, base_unit = to_base(m.group("amt"), m.group("unit"))
        if base_amt is None: return (None, None)
        return (int(base_amt) if float(base_amt).is_integer() else base_amt, base_unit)

    return (None, None)

def _clean_id(raw: Any) -> str:
    """Clean an ID/link cell: strip, drop NaN, drop trailing '.0' on float reads."""
    if raw is None or (not isinstance(raw, str) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if s.lower() == "nan":
        return ""
    if "." in s and s.replace(".", "").replace("-", "").isdigit():
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    return s

# ---------------------------------------------------------------------------
#  ZEPTO SELENIUM HELPERS (from the unified scraper)
# ---------------------------------------------------------------------------

def _parse_backup_links(raw) -> list:
    """Parse the 'Backup_Links' cell into a clean list of URL strings.

    Handles a Python-list-looking string (e.g. "['https://…', 'https://…']"),
    plain comma/newline/semicolon-separated URLs, or an actual list. Blanks
    dropped, original order preserved. ast.literal_eval is safe (literals only)."""
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(u).strip() for u in raw if str(u).strip()]
    s = str(raw).strip()
    if not s or s.lower() == "nan" or s in ("[]", "()"):
        return []
    if s.startswith("[") or s.startswith("("):
        try:
            import ast
            parsed = ast.literal_eval(s)
            if isinstance(parsed, (list, tuple)):
                return [str(u).strip() for u in parsed if str(u).strip()]
        except (ValueError, SyntaxError):
            pass
    s = s.strip("[]()").replace("'", "").replace('"', "")
    parts = re.split(r"[,\n;]+", s)
    return [p.strip() for p in parts if p.strip()]


def _qty_distance(target, candidate, target_unit: str = "", candidate_unit: str = ""):
    """Absolute numeric difference between target and candidate quantities, or
    None if either side isn't a plain number. Units are informational (distance
    is applied numerically even if units differ); only a non-numeric/composite
    side (e.g. "1000 g + 500 ml") is disqualified."""
    try:
        return abs(float(target) - float(candidate))
    except (ValueError, TypeError):
        return None


def _compute_final_qty_unit(weight_str: str) -> tuple:
    """Zepto Net-Qty string -> (final_qty, final_unit) with blank-string fallback,
    using the Zepto-specific parser (handles '500 g x 2', '10 N (1 kg)', etc.)."""
    try:
        qty_val, unit_val = _parse_weight_to_base_zepto(weight_str or "")
    except Exception:
        qty_val, unit_val = None, ""
    final_qty = qty_val if qty_val is not None else ""
    final_unit = unit_val if (qty_val is not None and unit_val) else ""
    return final_qty, final_unit


def _zpt_extract_price(text):
    try:
        if text is None: return None
        if isinstance(text, (int, float)): return int(text)
        m = re.search(r"₹[\d,.]+", str(text))
        if m:
            return int(m.group().replace("₹", "").replace(",", "").strip())
        digits = re.sub(r'[^\d.]', '', str(text))
        if digits:
            if '.' in digits:
                val = float(digits)
                return int(val) if val.is_integer() else int(val)
            return int(digits)
    except Exception:
        pass
    return None

def _zpt_is_valid_scrape(d: dict) -> bool:
    name = str(d.get("ProductName", "")).strip().lower()
    price = d.get("Price"); mrp = d.get("MRP")
    for pat in ["502 bad gateway", "404", "page not found", "error",
                "access denied", "forbidden", "503 service", "server error"]:
        if pat in name:
            return False
    if not name:
        return False
    if price is None and mrp is None:
        return False
    return True

def _zpt_click_select_location(driver, wait):
    """Open Zepto's address modal, handling BOTH header states:
      - no location set yet  -> the "Select Location" button
      - a location already set -> the header address chip, which is a
        <button aria-haspopup="dialog"> wrapping <h3 data-testid="user-address">
        (its class is obfuscated, so we key off the stable testid / aria attrs).
    Polls briefly for the header to render and clicks the first control found."""
    selectors = [
        '//button[@aria-label="Select Location"]',                                  # unset
        '//button[@aria-haspopup="dialog" and .//*[@data-testid="user-address"]]',   # set: address chip
        '//*[@data-testid="user-address"]/ancestor::button[1]',                      # set: chip via ancestor
        '//*[@data-testid="user-address"]',                                          # set: the chip itself
    ]
    for _ in range(12):   # up to ~6s for the header to render
        for sel in selectors:
            try:
                for el in driver.find_elements(By.XPATH, sel):
                    if el.is_displayed():
                        driver.execute_script("arguments[0].click();", el)
                        return True
            except Exception:
                continue
        time.sleep(0.5)
    return False

def _zpt_enter_pincode(driver, wait, pincode):
    try:
        inp = wait.until(EC.visibility_of_element_located(
            (By.XPATH, '//input[@placeholder="Search a new address"]')))
        inp.clear(); inp.send_keys(pincode)

        # Click the first address-search result (data-testid="address-search-item")
        try:
            first_result = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, '[data-testid="address-search-item"]')))
            driver.execute_script("arguments[0].click();", first_result)
            print(f"    [ZPT] clicked first address-search-item result")
        except Exception:
            # Fallback: press Enter to select whatever is highlighted
            print(f"    [ZPT] address-search-item not found, falling back to ENTER")
            inp.send_keys(Keys.ENTER)

        # Click "Confirm & Continue" if the confirmation panel appears.
        # Primary selector: data-testid="location-confirm-btn"
        # Fallback: aria-label="Confirm Action"
        try:
            confirm = wait.until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, '[data-testid="location-confirm-btn"]')))
            driver.execute_script("arguments[0].click();", confirm)
            print(f"    [ZPT] clicked Confirm & Continue")
        except Exception:
            try:
                confirm = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, '//button[@aria-label="Confirm Action"]')))
                driver.execute_script("arguments[0].click();", confirm)
                print(f"    [ZPT] clicked Confirm Action (aria-label fallback)")
            except Exception:
                print(f"    [ZPT] no Confirm & Continue button found — continuing anyway")
        return True
    except Exception:
        traceback.print_exc()
        return False

def _zpt_extract_store_id(driver) -> str:
    """Extract store_id from window.event_meta_string injected by Zepto's SSR."""
    try:
        raw = driver.execute_script(
            "var m = window.event_meta_string; "
            "if (!m) return ''; "
            "try { return JSON.parse(m).store_id || ''; } catch(e) { return ''; }"
        )
        if raw:
            return str(raw).strip()
    except Exception:
        pass
    # Fallback: regex over the page source in case the JS var isn't set yet
    try:
        src = driver.page_source
        m = re.search(r'"store_id"\s*:\s*"([a-f0-9\-]{36})"', src)
        if m:
            return m.group(1)
    except Exception:
        pass
    return ""

def _zpt_set_location(driver, wait, pincode, auto: bool) -> str:
    """Navigate to Zepto, set location for pincode, and return the store_id.

    Always fully automated — the `auto` parameter is kept for API compatibility
    but the manual ENTER prompt has been removed. Location selection now uses
    the real data-testid selectors from Zepto's page HTML."""
    driver.get("https://www.zepto.com/")
    time.sleep(3)

    located = False
    if _zpt_click_select_location(driver, wait):
        time.sleep(1.5)
        located = _zpt_enter_pincode(driver, wait, pincode)
        time.sleep(3)
    else:
        print(f"  [ZPT] could not open the location modal for {pincode} "
              f"(neither 'Select Location' nor the address chip was found)")

    if not located:
        print(f"  [ZPT {pincode}] location set may have failed — proceeding anyway")
    else:
        print(f"  [ZPT {pincode}] location confirmed automatically")

    # Allow page to settle then read the store_id Zepto writes into window globals
    time.sleep(2)
    store_id = _zpt_extract_store_id(driver)
    if store_id:
        print(f"  [ZPT {pincode}] store_id={store_id}")
    else:
        print(f"  [ZPT {pincode}] store_id not found in page globals")
    return store_id

def _zpt_scrape_page(driver, wait, url) -> dict:
    result = {k: "" for k in SCRAPED_FIELDS["zepto"]}
    driver.get(url); time.sleep(2)

    # Store_ID — read from window.event_meta_string on every product page
    result["Store_ID"] = _zpt_extract_store_id(driver)

    # ProductName + stock
    try:
        elems = driver.find_elements(By.CSS_SELECTOR, "#product-features-wrapper h1")
        name = ""
        for e in elems:
            t = e.text.strip()
            if "out of stock" in t.lower() or "please try changing" in t.lower():
                continue
            name = t; break
        result["ProductName"] = name or driver.title.replace(" | Zepto", "").strip()

        # Stock: when the selected variant is OOS, Zepto shows a banner (NOT in the
        # title h1) like:
        #   <h3>Current selection is out of stock</h3>
        #   <p>Please try changing the selection</p>
        # Those exact phrases only appear for the main product's OOS state (not in
        # the "you might also like" carousel), so matching them on the rendered page
        # is a precise signal. The old code only scanned the title h1 and so
        # defaulted everything to In Stock.
        try:
            body_txt = (driver.find_element(By.TAG_NAME, "body").text or "").lower()
        except Exception:
            body_txt = ""
        if ("current selection is out of stock" in body_txt
                or "please try changing the selection" in body_txt):
            result["is_out_of_stock"] = "Out of Stock"
        elif result["ProductName"]:
            result["is_out_of_stock"] = "In Stock"
        else:
            result["is_out_of_stock"] = "Unknown"
    except Exception:
        result["ProductName"] = ""; result["is_out_of_stock"] = "Unknown"

    # Weight
    try:
        nq = driver.find_element(By.XPATH, '//span[contains(normalize-space(.),"Net Qty")]')
        result["Weight"] = nq.text.strip().split("Net Qty:")[-1].strip()
    except Exception:
        result["Weight"] = ""

    # Price
    try:
        rs = driver.find_element(
            By.XPATH,
            '//*[@id="product-features-wrapper"]//span[normalize-space()="₹" and not(contains(@class,"line-through"))]',
        )
        vs = rs.find_element(By.XPATH, 'following-sibling::span[1]')
        result["Price"] = _zpt_extract_price(f"{rs.text}{vs.text}")
    except Exception:
        try:
            ap = driver.find_element(
                By.XPATH,
                '//*[@id="product-features-wrapper"]//*[contains(normalize-space(.),"₹") and not(contains(@class,"line-through"))]',
            )
            result["Price"] = _zpt_extract_price(ap.text)
        except Exception:
            result["Price"] = None

    # MRP
    try:
        m = driver.find_element(
            By.XPATH,
            '//span[normalize-space(text())="MRP"]/following-sibling::span[contains(text(),"₹")]',
        )
        result["MRP"] = _zpt_extract_price(m.text)
    except Exception:
        try:
            mf = driver.find_element(
                By.XPATH,
                '//span[contains(@class,"line-through") and contains(text(),"₹")]',
            )
            result["MRP"] = _zpt_extract_price(mf.text)
        except Exception:
            result["MRP"] = result.get("Price")

    # Saver
    try:
        s = driver.find_element(
            By.XPATH,
            '//*[@id="product-features-wrapper"]//div[contains(text(),"Get it for")]/span',
        )
        result["ZeptoSaverPrice"] = _zpt_extract_price(s.text.strip())
    except Exception:
        result["ZeptoSaverPrice"] = None

    # Use Zepto-specific parser — handles "500 g x 2", "10 N (1 kg)",
    # "12 × 250 ml" correctly where the generic parser does not.
    qv, uv = _parse_weight_to_base_zepto(result.get("Weight", ""))
    result["Final_Qty"] = qv if qv is not None else ""
    result["Final_unit"] = uv if (qv is not None and uv) else ""
    return result

def _make_driver(headless: bool = False):
    """Create a Selenium Chrome driver (one per run; its location is re-set per
    pincode). webdriver-manager fetches a matching chromedriver automatically."""
    options = ChromeOptions()
    options.add_argument("--start-maximized")
    if headless:
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1366,900")
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()),
        options=options,
    )
    return driver


# ---------------------------------------------------------------------------
#  PROGRESS / RESUME / JSON OUTPUT  (shared with Blinkit/FK)
# ---------------------------------------------------------------------------

def progress_key(pincode: str) -> str:
    return str(pincode)

def save_progress(progress: dict, path: Optional[Path] = None) -> None:
    """Atomic save: write .tmp then replace."""
    path = path or PROGRESS_FILE
    tmp = Path(str(path) + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2, default=str)
    tmp.replace(path)

def load_progress(path: Optional[Path] = None) -> Optional[dict]:
    """Load progress JSON, falling back to the .tmp copy if the main file is
    corrupt (e.g. process killed mid-write)."""
    p = Path(path or PROGRESS_FILE)
    for candidate in (p, Path(str(p) + ".tmp")):
        if candidate.exists():
            try:
                with open(candidate, encoding="utf-8") as f:
                    return json.load(f)
            except json.JSONDecodeError:
                continue
    return None

def _latest_dated_file(suffix: str) -> Optional[Path]:
    """Most recent existing BASE_DIR/{YYYY-MM-DD}_{suffix} file, or None. ISO
    dates sort chronologically as plain strings, so a filename sort suffices."""
    matches = sorted(BASE_DIR.glob(
        f"[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]_{suffix}"
    ))
    return matches[-1] if matches else None

def _hydrate_product(product: dict, saved: dict) -> None:
    """Overlay previously-saved fields onto a fresh template row (used on
    resume so skipped/already-done rows keep their scraped values)."""
    for k, v in saved.items():
        product[k] = v

def _slim_products(products: list[dict]) -> list[dict]:
    """Keep only template + scraped columns; coerce NaN to empty string."""
    keep = set(TEMPLATE_COLUMNS + SCRAPED_COLUMNS)
    out = []
    for p in products:
        s = {k: v for k, v in p.items() if k in keep}
        for k, v in list(s.items()):
            if isinstance(v, float) and v != v:   # NaN
                s[k] = ""
        out.append(s)
    return out

def _save_pincode_progress(progress: dict, pincode: str, products: list[dict],
                           success_indices, errored_indices,
                           tier: str = "", city: str = "") -> None:
    """Persist one pincode's state into the progress JSON (atomic)."""
    pkey = progress_key(pincode)
    entry = {
        "pincode": pincode, "tier": tier, "city": city,
        "products": _slim_products(products),
        "success_indices": sorted(success_indices),
        "errored_indices": sorted(errored_indices),
        "completed": len(errored_indices) == 0 and len(success_indices) >= len(products),
    }
    progress.setdefault("results", {})[pkey] = entry
    save_progress(progress)

def write_product_data_json(pincode: str, tier: str, city: str,
                            products: list[dict],
                            path: Optional[Path] = None) -> Path:
    """Persist merged scraped data into a SINGLE combined JSON, keyed by pincode.

    Each call loads the existing combined file (falling back to its .tmp on
    corruption), updates this pincode's entry, and atomically rewrites the
    whole file — so a multi-pincode run accumulates into one file.
    """
    path = path or COMBINED_DATA_PATH
    pincodes: dict = {}
    for candidate in (path, Path(str(path) + ".tmp")):
        if candidate.exists():
            try:
                existing = json.loads(candidate.read_text(encoding="utf-8"))
                pincodes = existing.get("pincodes", {}) or {}
                break
            except Exception:
                continue

    rows = _slim_products(products)
    pincodes[str(pincode)] = {
        "pincode": pincode, "tier": tier, "city": city,
        "generated_at": datetime.datetime.now().isoformat(),
        "row_count": len(rows),
        "rows": rows,
    }
    payload = {
        "generated_at": datetime.datetime.now().isoformat(),
        "pincode_count": len(pincodes),
        "pincodes": pincodes,
    }
    tmp = Path(str(path) + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str),
                   encoding="utf-8")
    tmp.replace(path)
    return path



# ---------------------------------------------------------------------------
#  EXCEL / PINCODE INPUT
# ---------------------------------------------------------------------------

def read_products_from_excel(excel_path, sheet_name: str = "Zepto") -> list[dict]:
    """Read the Zepto sheet from the template Excel. The Zepto product URL is
    expected in the 'Product_Link' column (same convention as the Blinkit / FK
    scrapers)."""
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Template Excel not found: {excel_path}")

    df = pd.read_excel(excel_path, sheet_name=sheet_name)

    link_col = "Product_Link"
    if link_col not in df.columns:
        raise ValueError(
            f"Column '{link_col}' not found in sheet '{sheet_name}'. "
            f"Found columns: {list(df.columns)}"
        )

    df = df.dropna(subset=[link_col])
    df = df[df[link_col].astype(str).str.strip() != ""]
    print(f"\U0001f4c4 Read {len(df)} products from '{sheet_name}' sheet in {excel_path.name}")
    return df.to_dict(orient="records")


def read_pincodes(
    pincodes_path: str | Path = PINCODES_PATH,
    company: str = "All",
) -> pd.DataFrame:
    """
    Read Sheet1 of the SHARED pincodes_to_run.xlsx and filter to the rows
    for one company. The same Excel is used by every platform scraper
    (Blinkit, Zepto, Instamart, Flipkart Minutes, Amazon Now), and the
    'Company' column lets each scraper pick the rows it cares about.

    Expected columns: Pincodes, Tier, City, Lat, Lon, Company
        - 'Company' is optional. Blank/missing → treated as 'All'.
        - Pass company='All' to use the default rows (most common).
        - Pass company='VMM' (etc.) for a one-off client-specific run.

    Match is case-insensitive and whitespace-tolerant.
    """
    pincodes_path = Path(pincodes_path)
    if not pincodes_path.exists():
        raise FileNotFoundError(f"Pincodes file not found: {pincodes_path}")

    df = pd.read_excel(pincodes_path, sheet_name=0)  # Sheet1
    for col in ("Pincodes", "Tier", "City", "Lat", "Lon"):
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found in pincodes file. Found: {list(df.columns)}")

    df["Pincodes"] = df["Pincodes"].astype(str).str.strip()
    df["Pincodes"] = df["Pincodes"].apply(_clean_pincode_str)
    df["Lat"] = df["Lat"].astype(str).str.strip()
    df["Lon"] = df["Lon"].astype(str).str.strip()

    # Normalize Company column → default to "All"
    if "Company" in df.columns:
        df["Company"] = df["Company"].fillna("All").astype(str).str.strip()
        df.loc[df["Company"] == "", "Company"] = "All"
    else:
        df["Company"] = "All"

    # Filter to the requested company (case-insensitive)
    target = (company or "All").strip().lower()
    filtered = df[df["Company"].str.lower() == target].reset_index(drop=True)

    if filtered.empty:
        available = sorted(df["Company"].unique().tolist())
        raise ValueError(
            f"No pincodes found for Company={company!r}. "
            f"Available companies in {pincodes_path.name}: {available}"
        )

    print(
        f"📄 Read {len(filtered)} pincodes from {pincodes_path.name} "
        f"(filter: Company == {company!r}; total rows in file: {len(df)})"
    )
    return filtered

def read_pincode_map(pincodes_path: str | Path = PINCODES_PATH) -> dict:
    """
    Read the 'Map' sheet → dict keyed by primary pincode.

    Expected columns: Pincode, Mapped, City, Lat, Lon
    There can be multiple rows per primary pincode (one per fallback option).

    Returns:
        {
            "400051": [
                {"mapped": "400034", "city": "Mumbai",
                 "lat": "19.05961", "lon": "72.8552983"},
                {"mapped": "400055", "city": "Mumbai",
                 "lat": "19.0855",  "lon": "72.8421"},
                ...
            ],
            ...
        }

    Order of fallbacks within each list = order they appear in the sheet,
    which is the order they will be tried.

    Returns an empty dict if the 'Map' sheet does not exist.
    """
    pincodes_path = Path(pincodes_path)
    if not pincodes_path.exists():
        return {}

    try:
        df = pd.read_excel(pincodes_path, sheet_name="Map")
    except (ValueError, KeyError):
        # Sheet doesn't exist — fallback chain disabled
        print("ℹ️  No 'Map' sheet found in pincodes file. OOS fallback disabled.")
        return {}

    for col in ("Pincode", "Mapped", "City", "Lat", "Lon"):
        if col not in df.columns:
            raise ValueError(
                f"Column '{col}' not found in 'Map' sheet. Found: {list(df.columns)}"
            )

    df["Pincode"] = df["Pincode"].astype(str).str.strip().apply(_clean_pincode_str)
    df["Mapped"] = df["Mapped"].astype(str).str.strip().apply(_clean_pincode_str)
    df["Lat"] = df["Lat"].astype(str).str.strip()
    df["Lon"] = df["Lon"].astype(str).str.strip()

    pincode_map: dict[str, list[dict]] = {}
    for _, row in df.iterrows():
        primary = row["Pincode"]
        pincode_map.setdefault(primary, []).append({
            "mapped": row["Mapped"],
            "city": str(row.get("City", "")).strip(),
            "lat": row["Lat"],
            "lon": row["Lon"],
        })

    total_mappings = sum(len(v) for v in pincode_map.values())
    print(f"📄 Read {total_mappings} fallback mappings for {len(pincode_map)} primary pincodes")
    return pincode_map
# ---------------------------------------------------------------------------
#  SCRAPE ONE PINCODE  (two-pass: primary, then batched OOS fallback)
# ---------------------------------------------------------------------------

def scrape_for_pincode(
    products_template: list,
    pincode: str,
    driver,
    wait,
    delay: float = 2.0,
    fallback_pincodes=None,
    auto_zepto: bool = True,
    progress=None,
    tier: str = "",
    city: str = "",
) -> list:
    """
    Scrape all products for one pincode using the shared Selenium `driver`,
    mirroring the FK/Blinkit resume + error-tracking + batched-fallback design.

      Pass 1 — set the driver location to the primary pincode, scrape every
               product. Out-of-Stock products are queued for fallback; pages that
               don't yield a valid scrape are queued for retry.
      Pass 2 — for each mapped fallback pincode IN ORDER, re-point the SAME driver
               at that pincode and re-check the whole unresolved batch; anything
               In Stock there resolves (Pincode_Used = that fallback). Whatever is
               still unresolved moves to the next fallback.

    A product that is OOS everywhere keeps its primary OOS data (success). A page
    that never yields a valid scrape anywhere is recorded as errored (retried on
    --resume). Progress is flushed every SAVE_EVERY products and once at the end.
    """
    products = copy.deepcopy(products_template)
    fallback_pincodes = fallback_pincodes or []

    # \u2500\u2500 Resume state \u2500\u2500
    pkey = progress_key(pincode)
    existing = (progress or {}).get("results", {}).get(pkey, {})
    success_indices = set(existing.get("success_indices", []))
    errored_indices = set(existing.get("errored_indices", []))
    for i, saved in enumerate(existing.get("products", [])):
        if i < len(products):
            _hydrate_product(products[i], saved)
    skip_indices = success_indices.copy()
    errored_indices.clear()

    n_done = len(skip_indices)
    n_retry = len(existing.get("errored_indices", []))
    if n_done or n_retry:
        print(f"   \u21bb resume: {n_done} done, {n_retry} to retry, "
              f"{len(products) - n_done - n_retry} new")

    print(f"\n\U0001f680 Scraping {len(products)} products for pincode {pincode}")
    if fallback_pincodes:
        print(f"   \u21b3 {len(fallback_pincodes)} fallback pincodes available for OOS retries")

    # \u2500\u2500 Pass 1: primary pincode \u2500\u2500
    store_id = _zpt_set_location(driver, wait, pincode, auto=auto_zepto)

    page_urls: dict = {}
    oos_pending: list = []
    failed_pending: list = []

    processed = 0
    for idx, product in enumerate(products):
        for key in SCRAPED_COLUMNS:
            product.setdefault(key, "")
        if idx in skip_indices:
            continue

        url = _clean_id(product.get("Product_Link", ""))
        if not url:
            success_indices.add(idx)
            continue
        page_urls[idx] = url
        product["Used_Link"] = url   # default; a backup may replace it in Pass 3

        try:
            parsed = _zpt_scrape_page(driver, wait, url)
            product.update(parsed)
            product["Pincode_Used"] = str(pincode)
            stock = parsed.get("is_out_of_stock")
            if stock == "Out of Stock":
                success_indices.add(idx)          # have OOS data
                if fallback_pincodes:
                    oos_pending.append(idx)        # try to upgrade to In Stock
                print(f"    [ZPT] {idx+1}/{len(products)}: "
                      f"{str(parsed.get('ProductName',''))[:40]}  OUT OF STOCK")
            elif _zpt_is_valid_scrape(parsed):
                success_indices.add(idx)
                print(f"    [ZPT] {idx+1}/{len(products)}: "
                      f"{str(parsed.get('ProductName',''))[:40]}  "
                      f"P={parsed.get('Price')} M={parsed.get('MRP')}  Stock={stock}")
            else:
                failed_pending.append(idx)         # invalid page -> retry at fallback/resume
                print(f"    [ZPT] {idx+1}/{len(products)}: invalid page (will retry)")
        except Exception as e:
            print(f"    [ZPT] error on row {idx+1}: {e}")
            failed_pending.append(idx)

        processed += 1
        if progress is not None and processed % SAVE_EVERY == 0:
            _save_pincode_progress(progress, pincode, products,
                                   success_indices, errored_indices, tier=tier, city=city)
        if delay > 0:
            time.sleep(delay)

    # \u2500\u2500 Pass 2: batched OOS / failure fallback (re-point same driver) \u2500\u2500
    failed_set = set(failed_pending)
    remaining = list(dict.fromkeys(oos_pending + failed_pending))

    if remaining and fallback_pincodes:
        print(f"\n   \U0001f504 {len(remaining)} product(s) unresolved at {pincode} "
              f"(OOS or failed); checking fallback pincodes in order\u2026")
        for fb in fallback_pincodes:
            if not remaining:
                break
            fb_pin = _clean_pincode_str(fb.get("mapped", ""))
            if not fb_pin:
                continue
            print(f"   \u21aa fallback {fb_pin} ({fb.get('city','')}): "
                  f"re-checking {len(remaining)} product(s)")
            _zpt_set_location(driver, wait, fb_pin, auto=auto_zepto)
            still = []
            for idx in remaining:
                try:
                    fb_parsed = _zpt_scrape_page(driver, wait, page_urls[idx])
                except Exception as e:
                    print(f"      [ZPT] error re-checking row {idx+1} at {fb_pin}: {e}")
                    still.append(idx)
                    if delay > 0:
                        time.sleep(delay)
                    continue
                if fb_parsed.get("is_out_of_stock") == "In Stock" and _zpt_is_valid_scrape(fb_parsed):
                    products[idx].update(fb_parsed)
                    products[idx]["Pincode_Used"] = str(fb_pin)
                    success_indices.add(idx)
                    errored_indices.discard(idx)
                    print(f"      \u2705 In Stock at {fb_pin}: "
                          f"{str(fb_parsed.get('ProductName',''))[:40]}")
                else:
                    still.append(idx)
                if delay > 0:
                    time.sleep(delay)
            remaining = still

        for idx in remaining:
            if idx in failed_set:
                errored_indices.add(idx)          # never valid anywhere -> retry on resume
            # OOS-everywhere products keep primary OOS data (stay success)
        if remaining:
            n_failed_left = sum(1 for i in remaining if i in failed_set)
            print(f"   \u2139\ufe0f  {len(remaining)} still unresolved after all fallbacks "
                  f"({n_failed_left} failed \u2192 errored, "
                  f"{len(remaining) - n_failed_left} kept as primary OOS data)")
    else:
        # No fallbacks (or nothing to retry): failed pages are errored
        for idx in failed_pending:
            errored_indices.add(idx)

    # \u2500\u2500 Final_Qty / Final_unit from each product's final Weight (Zepto parser) \u2500\u2500
    for idx, product in enumerate(products):
        if idx in skip_indices:
            continue
        qv, uv = _parse_weight_to_base_zepto(product.get("Weight", ""))
        product["Final_Qty"] = qv if qv is not None else ""
        product["Final_unit"] = uv if (qv is not None and uv) else ""

    # \u2500\u2500 Pass 3: backup-link closest-qty selection \u2500\u2500
    # For any product whose Final_Qty/unit doesn't match its Base_Final target,
    # scrape each Backup_Links entry AT THE WINNING PINCODE (primary, or the OOS
    # fallback that resolved it) and keep whichever link's Final_Qty is closest
    # to Base_Final quantity. The current result is seeded, so a backup is chosen
    # only if STRICTLY closer (ties keep the current link). Same-pincode: this
    # only swaps which product variant/link is used. Jobs are grouped by winning
    # pincode so the shared driver re-points its location once per group.
    backup_jobs = {}   # win_pin -> list of (idx, links, base_qty, base_unit, pq, pu)
    for idx, product in enumerate(products):
        if idx in skip_indices:
            continue
        if not str(product.get("ProductName", "")).strip():
            continue
        links = _parse_backup_links(product.get("Backup_Links", ""))
        if not links:
            continue
        base_qty = product.get("Base_Final quantity", "")
        base_unit = str(product.get("Base_Final unit", "") or "").strip()
        pq = product.get("Final_Qty", "")
        pu = str(product.get("Final_unit", "") or "").strip()
        qty_mismatch = _is_mismatch(base_qty, pq)
        unit_mismatch = bool(base_unit and pu and base_unit.lower() != pu.lower())
        if qty_mismatch or unit_mismatch:
            win_pin = product.get("Pincode_Used", pincode)
            backup_jobs.setdefault(str(win_pin), []).append(
                (idx, links, base_qty, base_unit, pq, pu))

    if backup_jobs:
        total = sum(len(v) for v in backup_jobs.values())
        print(f"\n   \U0001f501 backup-link closest-qty check for {total} product(s) "
              f"with a Base_Final mismatch\u2026")
        for win_pin, jobs in backup_jobs.items():
            # Re-point the shared driver at this group's winning pincode once.
            _zpt_set_location(driver, wait, win_pin, auto=auto_zepto)
            for idx, links, base_qty, base_unit, pq, pu in jobs:
                product = products[idx]
                print(f"    \U0001f501 row {idx+1}: Base={base_qty} {base_unit} vs "
                      f"Final={pq} {pu} \u2014 trying {len(links)} backup(s) at {win_pin}")
                best_dist = _qty_distance(base_qty, pq, base_unit, pu)
                best = None   # (b_url, b_parsed, b_qty, b_unit)
                for b_idx, b_url in enumerate(links):
                    b_url = _clean_id(b_url)
                    if not b_url:
                        continue
                    if delay > 0:
                        time.sleep(delay)
                    try:
                        b_parsed = _zpt_scrape_page(driver, wait, b_url)
                    except Exception as e:
                        print(f"      \u274c backup#{b_idx+1}: scrape error: {e}")
                        continue
                    b_qty, b_unit = _compute_final_qty_unit(b_parsed.get("Weight", ""))
                    dist = _qty_distance(base_qty, b_qty, base_unit, b_unit)
                    print(f"      \u21aa backup#{b_idx+1}: Final_Qty={b_qty} {b_unit}  "
                          f"Stock={b_parsed.get('is_out_of_stock')}  distance={dist}")
                    if dist is not None and (best_dist is None or dist < best_dist):
                        best_dist = dist
                        best = (b_url, b_parsed, b_qty, b_unit)
                if best is not None:
                    b_url, b_parsed, b_qty, b_unit = best
                    product.update(b_parsed)
                    product["Final_Qty"] = b_qty
                    product["Final_unit"] = b_unit
                    product["Used_Link"] = b_url
                    product["Pincode_Used"] = str(win_pin)   # unchanged (same location)
                    print(f"      \U0001f3c6 backup chosen: {b_url[:55]}  "
                          f"(Final_Qty={b_qty} {b_unit}, distance={best_dist})")
                else:
                    print(f"      \u2713 primary still closest to Base_Final quantity")

    # \u2500\u2500 Final persist \u2500\u2500
    if progress is not None:
        extras = {"zepto_store_id": store_id} if store_id else None
        _save_pincode_progress(progress, pincode, products,
                               success_indices, errored_indices, tier=tier, city=city)
        write_product_data_json(pincode, tier, city, products)
    print(f"   \u2713 pincode {pincode} done: {len(success_indices)} ok, "
          f"{len(errored_indices)} err")
    return products


# ---------------------------------------------------------------------------
#  OUTPUT WORKBOOK  (shared with Blinkit/FK)
# ---------------------------------------------------------------------------

def _format_header_row(ws, num_cols: int):
    """Apply header formatting to the first row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def _auto_width(ws, max_width: int = 30):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)

def _apply_body_formatting(ws):
    """Apply body font and borders to all data rows."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

def build_output_excel(
    all_results: dict,       # {sheet_name: list[dict]}
    pincode_info: dict,      # {sheet_name: {"pincode": str, "tier": str, "city": str}}
    output_path: Path,
):
    """
    Build the final multi-sheet Excel:
      1. One sheet per pincode: "Tier_Pincode"
      2. "Summary" sheet: mismatches + zero-price flags
      3. "Price_Comparison" sheet: all products with per-pincode MRP & Price
    """
    from openpyxl import Workbook

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    all_issues = []   # collect issues for Summary
    ordered_cols = TEMPLATE_COLUMNS + SCRAPED_COLUMNS

    # ── Per-pincode sheets ────────────────────────────────────────────────
    for sheet_name, products in all_results.items():
        df = pd.DataFrame(products)

        # Enforce column order
        final_cols = [c for c in ordered_cols if c in df.columns]
        remaining = [c for c in df.columns if c not in final_cols and c != "__review_status__"]
        final_cols.extend(remaining)
        df = df[final_cols]

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Handle NaN
                if pd.isna(val) if not isinstance(val, str) else False:
                    cell.value = ""
                else:
                    cell.value = val

        # Column index lookup
        col_map = {name: idx for idx, name in enumerate(df.columns)}

        info = pincode_info.get(sheet_name, {})
        pc = info.get("pincode", sheet_name)

        # Highlight mismatches & zero prices
        for row_idx in range(2, ws.max_row + 1):
            natpl_col = col_map.get("National / PL")
            natpl = ws.cell(row=row_idx, column=natpl_col + 1).value if natpl_col is not None else ""
            # --- Qty mismatch: Base_Final quantity vs Final_Qty ---
            base_qty_col = col_map.get("Base_Final quantity")
            final_qty_col = col_map.get("Final_Qty")
            if base_qty_col is not None and final_qty_col is not None:
                base_val = ws.cell(row=row_idx, column=base_qty_col + 1).value
                final_val = ws.cell(row=row_idx, column=final_qty_col + 1).value
                if _is_mismatch(base_val, final_val):
                    ws.cell(row=row_idx, column=base_qty_col + 1).fill = MISMATCH_FILL
                    ws.cell(row=row_idx, column=final_qty_col + 1).fill = MISMATCH_FILL
                    item_name = ws.cell(row=row_idx, column=col_map.get("Item", 0) + 1).value or ""
                    all_issues.append({
                        "Sheet": sheet_name,
                        "Pincode": pc,
                        "Row": row_idx,
                        "Item": item_name,
                        "Issue_Type": "Qty Mismatch",
                        "National / PL": natpl,
                        "Base_Final_quantity": base_val,
                        "Final_Qty": final_val,
                        "Base_Final_unit": ws.cell(row=row_idx, column=col_map.get("Base_Final unit", 0) + 1).value,
                        "Final_unit": ws.cell(row=row_idx, column=col_map.get("Final_unit", 0) + 1).value,
                        "Price": ws.cell(row=row_idx, column=col_map.get("Price", 0) + 1).value,
                        "MRP": ws.cell(row=row_idx, column=col_map.get("MRP", 0) + 1).value,
                    })

            # --- Unit mismatch: Base_Final unit vs Final_unit ---
            base_unit_col = col_map.get("Base_Final unit")
            final_unit_col = col_map.get("Final_unit")
            if base_unit_col is not None and final_unit_col is not None:
                base_u = str(ws.cell(row=row_idx, column=base_unit_col + 1).value or "").strip().lower()
                final_u = str(ws.cell(row=row_idx, column=final_unit_col + 1).value or "").strip().lower()
                if base_u and final_u and base_u != final_u:
                    ws.cell(row=row_idx, column=base_unit_col + 1).fill = MISMATCH_FILL
                    ws.cell(row=row_idx, column=final_unit_col + 1).fill = MISMATCH_FILL
                    item_name = ws.cell(row=row_idx, column=col_map.get("Item", 0) + 1).value or ""
                    # Only add if not already added as qty mismatch for same row
                    already = any(
                        i["Sheet"] == sheet_name and i["Row"] == row_idx and i["Issue_Type"] == "Unit Mismatch"
                        for i in all_issues
                    )
                    if not already:
                        all_issues.append({
                            "Sheet": sheet_name,
                            "Pincode": pc,
                            "Row": row_idx,
                            "Item": item_name,
                            "Issue_Type": "Unit Mismatch",
                            "National / PL": natpl,
                            "Base_Final_quantity": ws.cell(row=row_idx, column=col_map.get("Base_Final quantity", 0) + 1).value,
                            "Final_Qty": ws.cell(row=row_idx, column=col_map.get("Final_Qty", 0) + 1).value,
                            "Base_Final_unit": base_u,
                            "Final_unit": final_u,
                            "Price": ws.cell(row=row_idx, column=col_map.get("Price", 0) + 1).value,
                            "MRP": ws.cell(row=row_idx, column=col_map.get("MRP", 0) + 1).value,
                        })

            # --- Zero Price / MRP ---
            price_col = col_map.get("Price")
            mrp_col = col_map.get("MRP")
            if price_col is not None:
                pval = ws.cell(row=row_idx, column=price_col + 1).value
                mval = ws.cell(row=row_idx, column=mrp_col + 1).value if mrp_col is not None else None
                # Check if product had a link (not a blank-URL row)
                link_col_idx = col_map.get("Product_Link")
                link_val = ws.cell(row=row_idx, column=link_col_idx + 1).value if link_col_idx is not None else ""
                link_str = str(link_val or "").strip()
                has_link = bool(link_str) and link_str.lower() != "nan"

                if has_link and (_is_zero_price(pval) or _is_zero_price(mval)):
                    if _is_zero_price(pval):
                        ws.cell(row=row_idx, column=price_col + 1).fill = ZERO_PRICE_FILL
                    if mrp_col is not None and _is_zero_price(mval):
                        ws.cell(row=row_idx, column=mrp_col + 1).fill = ZERO_PRICE_FILL
                    item_name = ws.cell(row=row_idx, column=col_map.get("Item", 0) + 1).value or ""
                    all_issues.append({
                        "Sheet": sheet_name,
                        "Pincode": pc,
                        "Row": row_idx,
                        "Item": item_name,
                        "Issue_Type": "Zero Price/MRP",
                        "National / PL": natpl,
                        "Base_Final_quantity": "",
                        "Final_Qty": "",
                        "Base_Final_unit": "",
                        "Final_unit": "",
                        "Price": pval,
                        "MRP": mval,
                    })

        _format_header_row(ws, ws.max_column)
        _apply_body_formatting(ws)
        _auto_width(ws)
        ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────
    _build_summary_sheet(wb, all_issues)

    # ── Price Comparison sheet ────────────────────────────────────────────
    _build_price_comparison_sheet(wb, all_results, pincode_info)

    # Save
    wb.save(output_path)
    print(f"\n💾 Final workbook saved → {output_path}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   Issues flagged: {len(all_issues)}")

def _is_mismatch(base_val, final_val) -> bool:
    """Check if Base_Final quantity and Final_Qty are numerically different."""
    if base_val is None or base_val == "" or final_val is None or final_val == "":
        return False
    try:
        b = float(base_val)
        f = float(final_val)
        return abs(b - f) > 0.01
    except (ValueError, TypeError):
        # If one is a combined string like "1000.0 g + 500.0 ml", flag it
        return str(base_val).strip() != str(final_val).strip()

def _is_zero_price(val) -> bool:
    """Check if a price/MRP value is zero or effectively zero."""
    if val is None or val == "":
        return False
    try:
        return float(val) == 0
    except (ValueError, TypeError):
        return False

def _build_summary_sheet(wb, all_issues: list[dict]):
    """Create a formatted Summary sheet with all flagged issues."""
    ws = wb.create_sheet(title="Summary")

    summary_headers = [
        "Sheet", "Pincode", "Row", "Item", "National / PL", "Issue_Type",
        "Base_Final_quantity", "Final_Qty", "Base_Final_unit", "Final_unit",
        "Price", "MRP",
    ]

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Zepto Scraper — Issue Summary (Generated {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')})"
    title_cell.font = Font(bold=True, size=13, name="Arial", color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Counts by type
    from collections import Counter
    type_counts = Counter(i["Issue_Type"] for i in all_issues)

    row_num = 3
    ws.cell(row=row_num, column=1, value="Issue Breakdown:").font = Font(bold=True, size=11, name="Arial")
    row_num += 1
    for issue_type, count in type_counts.items():
        fill = MISMATCH_FILL if "Mismatch" in issue_type else ZERO_PRICE_FILL
        ws.cell(row=row_num, column=1, value=issue_type).font = BODY_FONT
        ws.cell(row=row_num, column=1).fill = fill
        ws.cell(row=row_num, column=2, value=count).font = BODY_FONT
        row_num += 1

    ws.cell(row=row_num, column=1, value=f"Total Issues: {len(all_issues)}").font = Font(
        bold=True, size=11, name="Arial"
    )
    row_num += 2

    # Header row for detail table
    for col_idx, hdr in enumerate(summary_headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for issue in all_issues:
        row_num += 1
        for col_idx, hdr in enumerate(summary_headers, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=issue.get(hdr, ""))
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            # Color-code by issue type
            if issue.get("Issue_Type") == "Zero Price/MRP":
                cell.fill = ZERO_PRICE_FILL
            elif "Mismatch" in issue.get("Issue_Type", ""):
                cell.fill = MISMATCH_FILL

    _auto_width(ws)
    ws.freeze_panes = ws.cell(row=row_num - len(all_issues) + 1, column=1).coordinate if all_issues else "A2"

def _build_price_comparison_sheet(
    wb, all_results: dict, pincode_info: dict
):
    """
    Build a 'Price_Comparison' sheet:
      Item | Quantity | Segment | Grocery | National / PL | <Pincode>_MRP | <Pincode>_Price | ...
    Uses the first pincode's data as the row template.
    """
    ws = wb.create_sheet(title="Price_Comparison")

    base_cols = ["Item", "Quantity", "Segment", "Grocery", "National / PL"]
    sheet_names = list(all_results.keys())

    if not sheet_names:
        ws["A1"] = "No data available"
        return

    # Use the first result set to get the product list
    first_products = all_results[sheet_names[0]]
    n_products = len(first_products)

    # Build header
    headers = list(base_cols)
    for sn in sheet_names:
        info = pincode_info.get(sn, {})
        pc = info.get("pincode", sn)
        headers.append(f"{pc}_MRP")
        headers.append(f"{pc}_Price")

    for col_idx, hdr in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=hdr)

    # Write data
    for row_idx in range(n_products):
        excel_row = row_idx + 2

        # Base columns from first sheet
        product_0 = first_products[row_idx]
        for col_idx, col_name in enumerate(base_cols, 1):
            val = product_0.get(col_name, "")
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            ws.cell(row=excel_row, column=col_idx, value=val)

        # Per-pincode MRP & Price
        offset = len(base_cols) + 1
        for sn in sheet_names:
            products = all_results[sn]
            if row_idx < len(products):
                p = products[row_idx]
                mrp = p.get("MRP", "")
                price = p.get("Price", "")
            else:
                mrp, price = "", ""

            ws.cell(row=excel_row, column=offset, value=mrp if mrp is not None else "")
            ws.cell(row=excel_row, column=offset + 1, value=price if price is not None else "")
            offset += 2

    _format_header_row(ws, ws.max_column)
    _apply_body_formatting(ws)
    _auto_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
#  MAIN ORCHESTRATOR
# ---------------------------------------------------------------------------

def run_multi_pincode_scraper(
    template_path=TEMPLATE_PATH,
    pincodes_path=PINCODES_PATH,
    output_path=None,
    sheet_name: str = "Zepto",
    company: str = "All",
    delay: float = 2.0,
    headless: bool = False,
    auto_zepto: bool = True,
    resume: bool = False,
):
    """
    Read template + pincodes, scrape every pincode with a single shared Selenium
    driver (re-setting its location per pincode), apply batched OOS fallback, and
    write the final multi-sheet workbook. Path/output/resume handling mirrors the
    Blinkit and FK Minutes scrapers.
    """
    global PROGRESS_FILE, COMBINED_DATA_PATH
    template_path = Path(template_path)
    pincodes_path = Path(pincodes_path)

    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    if resume:
        latest = _latest_dated_file("zepto_progress.json")
        if latest is not None:
            date_str = latest.name.split("_", 1)[0]
    PROGRESS_FILE = BASE_DIR / f"{date_str}_zepto_progress.json"
    COMBINED_DATA_PATH = BASE_DIR / f"{date_str}_zepto_product_data.json"

    progress = (load_progress() if resume else None) or {}
    if resume:
        print(f"  \u21bb resume: loaded {len(progress.get('results', {}))} pincode(s) "
              f"from {PROGRESS_FILE.name}")

    if output_path is None:
        now = datetime.datetime.now()
        time_str = now.strftime("%H-%M-%S")
        suffix = "" if company.lower() == "all" else f"_{company}"
        output_path = BASE_DIR / f"Zepto{suffix}_{date_str}_{time_str}.xlsx"
    else:
        output_path = Path(output_path)

    products_template = read_products_from_excel(template_path, sheet_name)
    pincodes_df = read_pincodes(pincodes_path, company=company)
    pincode_map = read_pincode_map(pincodes_path)

    all_results = {}
    pincode_info = {}

    print(f"\n{'='*70}")
    print(f"  Zepto Multi-Pincode Scraper  |  Company filter: {company}")
    print(f"  Products: {len(products_template)}  |  Pincodes: {len(pincodes_df)}")
    print(f"{'='*70}")

    driver = None
    wait = None
    try:
        for _, row in pincodes_df.iterrows():
            pincode = str(row["Pincodes"]).strip()
            tier = str(row["Tier"]).strip()
            city = str(row["City"]).strip()
            sn = f"{tier}_{pincode}"[:31]
            fallbacks = pincode_map.get(pincode, [])

            entry = progress.get("results", {}).get(progress_key(pincode))
            if resume and entry and entry.get("completed"):
                print(f"\n  \u2713 Pincode {pincode} already completed \u2014 reusing saved data")
                products = copy.deepcopy(products_template)
                for i, saved in enumerate(entry.get("products", [])):
                    if i < len(products):
                        _hydrate_product(products[i], saved)
                all_results[sn] = products
                pincode_info[sn] = {"pincode": pincode, "tier": tier, "city": city}
                continue

            print(f"\n{'\u2500'*70}")
            print(f"  \U0001f4cc Pincode: {pincode}  |  Tier: {tier}  |  City: {city}  "
                  f"|  Fallbacks: {len(fallbacks)}")
            print(f"{'\u2500'*70}")

            # Lazily create the single shared driver on first pincode that needs it
            if driver is None:
                print("   \U0001f310 launching Chrome for Zepto\u2026")
                driver = _make_driver(headless)
                wait = WebDriverWait(driver, 15)

            products = scrape_for_pincode(
                products_template, pincode, driver, wait,
                delay=delay, fallback_pincodes=fallbacks, auto_zepto=auto_zepto,
                progress=progress, tier=tier, city=city,
            )
            all_results[sn] = products
            pincode_info[sn] = {"pincode": pincode, "tier": tier, "city": city}
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass

    build_output_excel(all_results, pincode_info, output_path)
    return output_path


# ---------------------------------------------------------------------------
#  ENTRY POINT
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(
        description="Zepto multi-pincode product scraper (Selenium-based)."
    )
    ap.add_argument("--resume", action="store_true",
                    help="Resume from the most recent dated progress JSON: skip "
                         "completed pincodes/rows, retry previously-errored rows.")
    ap.add_argument("--company", default="All",
                    help="Filter pincodes_to_run.xlsx rows by company (default: All).")
    ap.add_argument("--delay", type=float, default=2.0,
                    help="Delay in seconds between page loads (default: 2.0).")
    ap.add_argument("--sheet", default="Zepto",
                    help="Template sheet name (default: 'Zepto').")
    ap.add_argument("--headless", action="store_true",
                    help="Run Chrome headless.")
    ap.add_argument("--auto-zepto", action="store_true", default=True,
                    help="Automated location selection (default; kept for parity).")
    ap.add_argument("--output", default=None,
                    help="Explicit output .xlsx path (default: dated auto-name).")
    args = ap.parse_args()

    run_multi_pincode_scraper(
        output_path=args.output,
        sheet_name=args.sheet,
        company=args.company,
        delay=args.delay,
        headless=args.headless,
        auto_zepto=args.auto_zepto,
        resume=args.resume,
    )
